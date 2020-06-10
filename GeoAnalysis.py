# Analysis of geographical locations from processed data (分析基础)

import pandas as pd
from openpyxl import Workbook


def columndist(header):
    # get corresponding row from input Dataframe
    lstColumnRaw = dfInput[header]
    # generate distribution dictionary
    outputdict = dict()
    for i in lstColumnRaw:
        if i in outputdict:
            outputdict[i] += 1
        else:
            outputdict[i] = 1
    return outputdict


def dicttoexcel(dictname,sheetname,indexheader,valueheader):
    # Create sheet and write headers
    wsTemp = wbOutput.create_sheet(title=sheetname)
    wsTemp.cell(column=1, row=1, value=indexheader)
    wsTemp.cell(column=2, row=1, value=valueheader)
    # Write dictionary index vertically starting from row 2
    tempRow = 2
    for i in dictname:
        wsTemp.cell(column=1, row=tempRow, value=i)
        tempRow += 1
    # Write dictionary values vertically starting from row 2
    tempRow = 2
    for i in dictname:
        wsTemp.cell(column=2, row=tempRow, value=dictname[i])
        tempRow += 1


# Debug switch
isDebug = False

# Set input file (single file only) and create output workbook
dfInput = pd.read_excel("200608-报名原始数据.xls", read_only=True)
wbOutput = Workbook()

dictProvinceDist = columndist("省")
dicttoexcel(dictProvinceDist, "省份统计", "省份", "人数")

dictCompetitionDist = columndist("matchname")
dicttoexcel(dictCompetitionDist, "赛事统计", "赛事名称", "人数")

dictOrganizationDist = columndist("机构")
dicttoexcel(dictOrganizationDist, "机构统计", "机构名称", "人数")

# Save workbook
wbOutput.save(filename="200608-0-测试文件.xlsx")
