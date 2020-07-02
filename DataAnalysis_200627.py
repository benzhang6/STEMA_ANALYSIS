# Analysis of geographical locations from processed data

import pandas as pd
from openpyxl import Workbook


def columndist(header):
    # get corresponding row from input Dataframe
    lstColumnRaw = dfInput[header]
    # generate distribution dictionary
    outputdict = dict()
    for i in lstColumnRaw:
        if i in outputdict:
            outputdict[i] += 1
        else:
            outputdict[i] = 1
    lstsorted = sorted(outputdict.items(), key=lambda item: item[1], reverse=True)
    if isDebug:
        print(lstsorted)
    return lstsorted


def tuplelisttows(lstname, startingcolumn, startingrow, wsname):
    temprow1 = startingrow
    temprow2 = startingrow
    for i in lstname:
        wsname.cell(temprow1, startingcolumn, i[0])
        temprow1 += 1
        wsname.cell(temprow2, startingcolumn+1, i[1])
        temprow2 += 1


# Debug switch
isDebug = False

# Set input and output file
dfInput = pd.read_excel("200620-省赛考生分布.xlsx", read_only=True)
wbOutput = Workbook()
wsMain = wbOutput.active
wsMain.title = "机构统计"
wsCompetition = wbOutput.create_sheet(title="比赛名称统计")
wsProvince = wbOutput.create_sheet(title="省份统计")
wsGroup = wbOutput.create_sheet(title="组别统计")

wsMain.cell(1, 1, "所属机构")
wsMain.cell(1, 2, "人数")
wsCompetition.cell(1, 1, "比赛名称")
wsCompetition.cell(1, 2, "人数")

lstOrganization = columndist("所属机构")
tuplelisttows(lstOrganization, 1, 2, wsMain)

lstCompetition = columndist("比赛名称")
tuplelisttows(lstCompetition, 1, 2, wsCompetition)

dictProvince = {"河北": 0, "山西": 0, "辽宁": 0, "吉林": 0, "黑龙江": 0, "江苏": 0, "浙江": 0, "安徽": 0, "福建": 0, "江西": 0, "山东": 0,
                "河南": 0, "湖北": 0, "湖南": 0, "广东": 0, "海南": 0, "四川": 0, "贵州": 0, "云南": 0, "陕西": 0, "甘肃": 0, "青海": 0,
                "台湾": 0, "内蒙古": 0, "广西": 0, "西藏": 0, "宁夏": 0, "新疆": 0, "北京": 0, "上海": 0, "天津": 0, "重庆": 0, "香港": 0,
                "澳门": 0, "其他": 0}

for x in lstCompetition:
    boolFoundProvince = False
    for i in dictProvince:
        if x[0].find(i) != -1:
            dictProvince[i] += x[1]
            boolFoundProvince = True
    if not boolFoundProvince:
        dictProvince["其他"] += x[1]

if isDebug:
    print(dictProvince)

wsProvince.cell(1, 1, "省份")
wsProvince.cell(1, 2, "人数")

lstProvince = sorted(dictProvince.items(), key=lambda item: item[1], reverse=True)
if isDebug:
    print(lstProvince)
tuplelisttows(lstProvince, 1, 2, wsProvince)

dictGroup = {"Scratch": 0, "Arduino": 0, "C++": 0, "EV3": 0, "Python": 0, "Minecraft": 0, "其他": 0}

for x in lstCompetition:
    boolFoundGroup = False
    for i in dictGroup:
        if x[0].find(i) != -1:
            dictGroup[i] += x[1]
            boolFoundGroup = True
    if not boolFoundGroup:
        dictGroup["其他"] += x[1]

wsGroup.cell(1, 1, "组别")
wsGroup.cell(2, 1, "人数")

lstGroup = sorted(dictGroup.items(), key=lambda item: item[1], reverse=True)
tuplelisttows(lstGroup, 1, 2, wsGroup)


# Save workbook
wbOutput.save(filename="200620-0-测试文件.xlsx")
