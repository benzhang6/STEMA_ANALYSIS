# Calculates Avg. Score by Province from xxxxxx-4-发布成绩.xlsx

import pandas as pd
from openpyxl import Workbook


def avgscore(inputdf, header):
    lstScore = inputdf[header]
    lstProvince = inputdf["省份"]
    lstRawData = []
    for i in range(len(lstScore)):
        lstRawData.append((lstProvince[i], lstScore[i]))
    if isDebug:
        print("lstRawData of", header, ":", lstRawData)
    dictProvince = {"河北": [0, 0], "山西": [0, 0], "辽宁": [0, 0], "吉林": [0, 0], "黑龙江": [0, 0], "江苏": [0, 0], "浙江": [0, 0],
                    "安徽": [0, 0], "福建": [0, 0], "江西": [0, 0], "山东": [0, 0], "河南": [0, 0], "湖北": [0, 0], "湖南": [0, 0],
                    "广东": [0, 0], "海南": [0, 0], "四川": [0, 0], "贵州": [0, 0], "云南": [0, 0], "陕西": [0, 0], "甘肃": [0, 0],
                    "青海": [0, 0], "台湾": [0, 0], "内蒙古": [0, 0], "广西": [0, 0], "西藏": [0, 0], "宁夏": [0, 0], "新疆": [0, 0],
                    "北京": [0, 0], "上海": [0, 0], "天津": [0, 0], "重庆": [0, 0], "香港": [0, 0], "澳门": [0, 0], "其他": [0, 0]}
    for i in lstRawData:
        dictProvince[i[0]][0] += i[1]
        dictProvince[i[0]][1] += 1
    if isDebug:
        print("dictProvince of", header, ":", dictProvince)
    dictAvgScore = {"河北": 0, "山西": 0, "辽宁": 0, "吉林": 0, "黑龙江": 0, "江苏": 0, "浙江": 0, "安徽": 0, "福建": 0, "江西": 0, "山东": 0,
                    "河南": 0, "湖北": 0, "湖南": 0, "广东": 0, "海南": 0, "四川": 0, "贵州": 0, "云南": 0, "陕西": 0, "甘肃": 0, "青海": 0,
                    "台湾": 0, "内蒙古": 0, "广西": 0, "西藏": 0, "宁夏": 0, "新疆": 0, "北京": 0, "上海": 0, "天津": 0, "重庆": 0, "香港": 0,
                    "澳门": 0, "其他": 0}
    for i in dictProvince:
        if dictProvince[i][1] == 0:
            dictAvgScore[i] = 0
        else:
            dictAvgScore[i] = round(dictProvince[i][0] / dictProvince[i][1], 1)
    if isDebug:
        print("dictAvgScore of", header, ":", dictAvgScore)
    return dictAvgScore


def provincedist(inputdf):
    lstProvince = inputdf["省份"]
    dictProvince = {"河北": 0, "山西": 0, "辽宁": 0, "吉林": 0, "黑龙江": 0, "江苏": 0, "浙江": 0, "安徽": 0, "福建": 0, "江西": 0, "山东": 0,
                    "河南": 0, "湖北": 0, "湖南": 0, "广东": 0, "海南": 0, "四川": 0, "贵州": 0, "云南": 0, "陕西": 0, "甘肃": 0, "青海": 0,
                    "台湾": 0, "内蒙古": 0, "广西": 0, "西藏": 0, "宁夏": 0, "新疆": 0, "北京": 0, "上海": 0, "天津": 0, "重庆": 0, "香港": 0,
                    "澳门": 0, "其他": 0}
    for i in lstProvince:
        dictProvince[i] += 1
    return dictProvince


def dictvaluetorows(dictname, startingrow, startingcolumn):
    tempRow = startingrow
    for index in dictname:
        wsMain.cell(tempRow, startingcolumn, value=dictname[index])
        tempRow += 1


def dictindextorows(dictname, startingrow, startingcolumn):
    tempRow = startingrow
    for index in dictname:
        wsMain.cell(tempRow, startingcolumn, value=index)
        tempRow += 1


def scoreaddup(column1, column2, outputcolumn):
    for i in range(35):
            wsMain.cell(i + 2, outputcolumn, value=(wsMain.cell(i + 2, column1).value + wsMain.cell(i + 2, column2).value))


isDebug = False

dfInput200530 = pd.read_excel("200530-4-发布成绩.xlsx", read_only = True)
dfInput200112 = pd.read_excel("200112-4-发布成绩.xlsx", read_only = True)
dfInput191215 = pd.read_excel("191215-4-发布成绩.xlsx", read_only = True)
wbOutput = Workbook()
wsMain = wbOutput.active
wsMain.title = "省份平均分统计"
wsMain.cell(1, 1, value="省份")
wsMain.cell(1, 2, value="191215_人数")
wsMain.cell(1, 3, value="191215_第一部分平均分")
wsMain.cell(1, 4, value="191215_第二部分平均分")
wsMain.cell(1, 5, value="191215_总成绩平均分")
wsMain.cell(1, 7, value="200112_人数")
wsMain.cell(1, 8, value="200112_第一部分平均分")
wsMain.cell(1, 9, value="200112_第二部分平均分")
wsMain.cell(1, 10, value="200112_总成绩平均分")
wsMain.cell(1, 12, value="200530_人数")
wsMain.cell(1, 13, value="200530_第一部分平均分")
wsMain.cell(1, 14, value="200530_第二部分平均分")
wsMain.cell(1, 15, value="200530_总成绩平均分")

dict191215_p1 = avgscore(dfInput191215, "第一部分成绩")
dict191215_p2 = avgscore(dfInput191215, "第二部分成绩")
dictProvince191215 = provincedist(dfInput191215)

dict200112_p1 = avgscore(dfInput200112, "第一部分成绩")
dict200112_p2 = avgscore(dfInput200112, "第二部分成绩")
dictProvince200112 = provincedist(dfInput200112)

dict200530_p1 = avgscore(dfInput200530, "第一部分成绩")
dict200530_p2 = avgscore(dfInput200530, "第二部分成绩")
dictProvince200530 = provincedist(dfInput200530)

dictindextorows(dict191215_p1, 2, 1)
dictvaluetorows(dictProvince191215, 2, 2)
dictvaluetorows(dict191215_p1, 2, 3)
dictvaluetorows(dict191215_p2, 2, 4)
scoreaddup(3, 4, 5)
dictvaluetorows(dictProvince200112, 2, 7)
dictvaluetorows(dict200112_p1, 2, 8)
dictvaluetorows(dict200112_p2, 2, 9)
scoreaddup(8, 9, 10)
dictvaluetorows(dictProvince200530, 2, 12)
dictvaluetorows(dict200530_p1, 2, 13)
dictvaluetorows(dict200530_p2, 2, 14)
scoreaddup(13, 14, 15)


wbOutput.save(filename="200530-9-省份平均分统计.xlsx")
